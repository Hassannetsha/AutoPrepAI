import io
from pathlib import Path

import openpyxl
import pandas as pd
import pytest

from utils.excel_utils import read_excel_clean


class TestReadExcelClean:
    def test_sample_data_cleans_metadata_rows_and_empty_cols(self):
        df = read_excel_clean(
            Path(__file__).parents[1] / "Input" / "Project-Management-Sample-Data.xlsx"
        )
        assert df.shape == (41, 6)
        assert list(df.columns) == [
            "Project Name",
            "Task Name",
            "Assigned to",
            "Start Date",
            "Days Required",
            "End Date",
        ]
        assert not any("Unnamed" in c for c in df.columns)

    def test_clean_excel_no_metadata(self):
        buf = io.BytesIO()
        pd.DataFrame({"A": [1, 2], "B": [3, 4]}).to_excel(buf, index=False)
        buf.seek(0)
        df = read_excel_clean(buf)
        assert df.shape == (2, 2)
        assert list(df.columns) == ["A", "B"]

    def test_empty_excel_returns_empty(self):
        buf = io.BytesIO()
        pd.DataFrame().to_excel(buf, index=False)
        buf.seek(0)
        df = read_excel_clean(buf)
        assert df.empty

    def test_only_metadata_rows_returns_empty(self):
        buf = io.BytesIO()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(row=1, column=2, value="Title Row")
        wb.save(buf)
        buf.seek(0)
        df = read_excel_clean(buf)
        # Single cell gets detected as a header (1 col → min_header_cols=1),
        # and there are no data rows below it.
        assert df.empty

    def test_bytesio_input(self):
        with open(
            Path(__file__).parents[1] / "Input" / "Project-Management-Sample-Data.xlsx",
            "rb",
        ) as f:
            df = read_excel_clean(io.BytesIO(f.read()))
        assert df.shape == (41, 6)

    def test_bytes_input(self):
        with open(
            Path(__file__).parents[1] / "Input" / "Project-Management-Sample-Data.xlsx",
            "rb",
        ) as f:
            df = read_excel_clean(f.read())
        assert df.shape == (41, 6)

    def test_single_column(self):
        buf = io.BytesIO()
        pd.DataFrame({"x": [10, 20, 30]}).to_excel(buf, index=False)
        buf.seek(0)
        df = read_excel_clean(buf)
        assert df.shape == (3, 1)
        assert list(df.columns) == ["x"]

    def test_single_row(self):
        buf = io.BytesIO()
        pd.DataFrame({"a": [1], "b": [2]}).to_excel(buf, index=False)
        buf.seek(0)
        df = read_excel_clean(buf)
        assert df.shape == (1, 2)

    def test_trailing_empty_rows_and_cols_removed(self):
        buf = io.BytesIO()
        raw = pd.DataFrame({"A": [1, 2], "B": [3, 4]})
        with pd.ExcelWriter(buf, engine="openpyxl") as writer:
            raw.to_excel(writer, index=False, startrow=3, startcol=2)
        buf.seek(0)
        df = read_excel_clean(buf)
        assert df.shape == (2, 2)
        assert list(df.columns) == ["A", "B"]

    def test_duplicate_column_names_deduplicated(self):
        buf = io.BytesIO()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["x", "x", "y"])
        ws.append([1, 2, 3])
        wb.save(buf)
        buf.seek(0)
        df = read_excel_clean(buf)
        assert list(df.columns) == ["x", "x_1", "y"]

    def test_header_name_stripped(self):
        buf = io.BytesIO()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="  spaced  ")
        ws.cell(row=1, column=2, value="normal")
        ws.cell(row=2, column=1, value=1)
        ws.cell(row=2, column=2, value=2)
        wb.save(buf)
        buf.seek(0)
        df = read_excel_clean(buf)
        assert list(df.columns) == ["spaced", "normal"]

    def test_string_values_stripped(self):
        buf = io.BytesIO()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["name"])
        ws.append(["  hello  "])
        ws.append(["  world  "])
        wb.save(buf)
        buf.seek(0)
        df = read_excel_clean(buf)
        assert df["name"].tolist() == ["hello", "world"]

    def test_mixed_types_preserved(self):
        buf = io.BytesIO()
        raw = pd.DataFrame({
            "text": ["foo"],
            "number": [42],
            "date": [pd.Timestamp("2024-01-01")],
        })
        raw.to_excel(buf, index=False)
        buf.seek(0)
        df = read_excel_clean(buf)
        assert df["text"].iloc[0] == "foo"
        assert df["number"].iloc[0] == 42
        assert str(df["date"].iloc[0])[:10] == "2024-01-01"

    def test_path_string_input(self):
        path = str(
            Path(__file__).parents[1] / "Input" / "Project-Management-Sample-Data.xlsx"
        )
        df = read_excel_clean(path)
        assert df.shape == (41, 6)
